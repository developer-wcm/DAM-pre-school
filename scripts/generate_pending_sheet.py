"""Generate a 'Pending Work' Excel report for the DAM PreSchool app.

Items are reconciled against the actual code as of 2026-06-16, not just the
stale docs/PROGRESS_TRACKING.md (dated 2026-05-25).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (Area, Item, Status, Priority, Source, Notes / next step)
# Status: Pending | In progress | Needs verification
ROWS = [
    # --- Partially implemented (from code + progress doc, reconciled) ---
    ("Communication", "Realtime announcements (parent view)", "In progress", "High",
     "PROGRESS_TRACKING.md / README_DASHBOARDS.md",
     "Parents see announcements but no Supabase realtime subscription yet."),
    ("Communication", "Parent–Teacher chat", "Pending", "High",
     "PROGRESS_TRACKING.md",
     "Messaging UI exists for admin; parent<->teacher channel not connected."),
    ("Communication", "Unread message indicators", "Pending", "Medium",
     "README_DASHBOARDS.md", "No unread badges on conversations."),
    ("Communication", "File attachments in messages/announcements", "Pending", "Low",
     "README_DASHBOARDS.md", "No file sharing support yet."),

    # --- Fees / payments ---
    ("Fees", "Online fee payment integration", "Pending", "High",
     "PROGRESS_TRACKING.md", "Fees are view-only; no payment gateway wired."),
    ("Fees", "Admin fee collection / processing", "Pending", "High",
     "PROGRESS_TRACKING.md", "No way to record/collect payments from admin side."),
    ("Fees", "Payment receipt PDF download", "Pending", "Medium",
     "docs/PAYMENT_RECEIPT_FEATURE.md", "Receipt screen exists; 'Download PDF' marked coming soon."),

    # --- Appointments ---
    ("Appointments", "Reschedule request backend", "Pending", "High",
     ".kiro/specs/parent-appointment-modals (task 8.8)",
     "Reschedule details only logged; replace with real backend call."),
    ("Appointments", "Update appointment in backend on reschedule", "Pending", "High",
     ".kiro/specs/.../design.md (TODO)", "Backend mutation not implemented."),

    # --- Admin / management ---
    ("Admin", "Class assignment (assign teachers to classes)", "Pending", "Medium",
     "PROGRESS_TRACKING.md", "Admin cannot assign teachers to classes."),
    ("Admin", "Holiday management (create/delete)", "Needs verification", "Medium",
     "PROGRESS_TRACKING.md vs schema", "Doc says view-only; holidays table + RLS exist — confirm admin CRUD UI."),
    ("Admin", "Reports analytics / charts", "Pending", "Medium",
     "PROGRESS_TRACKING.md", "Reports screen present; charts not connected to data."),
    ("Admin", "Admin profile editing", "Pending", "Low",
     "app/(dashboard)/admin-profile.tsx:169", "'Coming Soon' alert on edit profile."),
    ("Admin", "'More' menu unrouted items", "Pending", "Low",
     "app/(dashboard)/more.tsx:163", "Menu items without a route show 'Coming Soon'."),

    # --- Teacher ---
    ("Teacher", "Teacher student list from DB", "Needs verification", "Medium",
     "PROGRESS_TRACKING.md", "Doc says not fetched from DB; verify against current teacher screens."),

    # --- Routing / known bugs ---
    ("Routing", "Principal routes to admin dashboard", "Needs verification", "High",
     "PROGRESS_TRACKING.md (known issue)",
     "Doc flags principal landing on admin dashboard; verify auth-routing.ts handles principal."),

    # --- Notifications (reconciled: now implemented) ---
    ("Notifications", "Push notifications end-to-end", "Needs verification", "High",
     "code: usePushNotifications.ts / send-push",
     "IMPLEMENTED since the doc was written. Verify delivery on a physical prod build via adb logcat."),

    # --- Auth (recent work) ---
    ("Auth", "Google OAuth PKCE fix", "Needs verification", "High",
     "lib/supabase.ts / context/auth.tsx (this session)",
     "flowType:pkce + Linking.createURL added. Rebuild APK and confirm sign-in completes."),

    # --- Infra / quality ---
    ("Infra", "Consolidate SQL scripts into migrations", "Pending", "Medium",
     "supabase/*.sql", "Dozens of ad-hoc fix_/debug_ scripts; no versioned migrations folder."),
    ("Infra", "Error tracking service (Sentry/Bugsnag)", "Pending", "Low",
     "utils/errorHandler.ts:146", "TODO: send errors to a tracking service."),
    ("Quality", "Offline support / caching", "Pending", "Low",
     "README_DASHBOARDS.md", "No offline caching for data fetches."),
    ("Quality", "Export reports (PDF/Excel)", "Pending", "Low",
     "README_DASHBOARDS.md", "Report export not implemented."),
]

STATUS_FILL = {
    "Pending": "FFF4CCCC",          # light red
    "In progress": "FFFFF2CC",      # light amber
    "Needs verification": "FFD9E1F2" # light blue
}
PRIORITY_FILL = {
    "High": "FFE06666",
    "Medium": "FFFFD966",
    "Low": "FFB6D7A8",
}

wb = Workbook()
ws = wb.active
ws.title = "Pending Work"

headers = ["#", "Area", "Item", "Status", "Priority", "Source", "Notes / Next step"]
header_fill = PatternFill("solid", fgColor="FF1B3A6B")  # school navy
header_font = Font(bold=True, color="FFFFFFFF", size=11)
thin = Side(style="thin", color="FFBFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title row
ws.merge_cells("A1:G1")
t = ws["A1"]
t.value = "DAM PreSchool — Pending Work  (generated 2026-06-16, reconciled with code)"
t.font = Font(bold=True, size=14, color="FF1B3A6B")
t.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 26

# Header row
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=2, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[2].height = 22

# Data rows
for i, (area, item, status, prio, source, notes) in enumerate(ROWS, start=1):
    r = i + 2
    values = [i, area, item, status, prio, source, notes]
    for col, v in enumerate(values, start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal="center" if col in (1, 4, 5) else "left")
    ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=STATUS_FILL.get(status, "FFFFFFFF"))
    ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=PRIORITY_FILL.get(prio, "FFFFFFFF"))

# Column widths
widths = [4, 16, 38, 18, 10, 34, 52]
for col, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col)].width = w

# Freeze header, add autofilter
ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:G{len(ROWS) + 2}"

out = "DAM_PreSchool_Pending_Work.xlsx"
wb.save(out)
print(f"Wrote {out} with {len(ROWS)} items")
